"""FileHandler Class"""

import csv
import io
import json
import os
from abc import ABC

import openpyxl
import pandas as pd
import yaml
from pyspark.sql import DataFrame, SparkSession
from pyspark.sql.types import StringType, StructField, StructType, TimestampType

def read_csv_txt_to_dataframe(
    spark: SparkSession, filepath: str, separator: str = ";"
) -> DataFrame:
    # split of as a separate function to prevent code duplication (sonarqube)
    df = spark.read.load(
        filepath,
        format="csv",
        sep=separator,
        inferSchema=False,
        header=True,
    )

    return df


class FileHandler:
    """File handler that reads different file formats based on strategy"""

    def __init__(self, strategy):
        self.strategy = strategy

    def read(self, filepath: str):
        return self.strategy.read(filepath)

    def read_to_dataframe(
        self, spark: SparkSession, filepath: str, **kwargs
    ) -> DataFrame:
        return self.strategy.read_to_dataframe(spark, filepath, **kwargs)


class FileHandlerContext(ABC):
    @classmethod
    def factory(cls, filepath: str) -> FileHandler:
        _, file_extension = os.path.splitext(filepath.lower())

        factory = {
            ".yml": YamlHandler(),
            ".json": JsonHandler(),
            ".csv": CsvHandler(),
            ".txt": TxtHandler(),
            ".xlsx": XlsxHandler(),
            "": ParquetHandler(),
        }

        if file_extension in factory.keys():
            file_handler = FileHandler(factory[file_extension])
            return file_handler

        raise NotImplementedError(
            f"File extension `{file_extension}` is not supported."
        )


class FileHandlerStrategy:
    def read(self, filepath: str):
        pass

    def read_to_dataframe(self, spark: SparkSession, filepath: str):
        pass


class YamlHandler(FileHandlerStrategy):
    def read(self, filepath):
        with open(filepath, "r") as file:
            return yaml.safe_load(file)

    def read_to_dataframe(self, spark: SparkSession, filepath: str):
        data_dict = self.read(filepath)
        df = spark.createDataFrame(data_dict.items())
        return df


class JsonHandler(FileHandlerStrategy):
    def read(self, filepath):
        with open(filepath, "r") as file:
            return json.load(file)

    def read_to_dataframe(self, spark: SparkSession, filepath: str) -> DataFrame:
        df = spark.read.json(filepath)
        return df


class CsvHandler(FileHandlerStrategy):
    def read(self, filepath: str):
        with open(filepath, "r") as file:
            csvFile = csv.reader(file)
            return list(csvFile)

    def read_to_dataframe(
        self, spark: SparkSession, filepath: str, separator: str = ";"
    ) -> DataFrame:
        df = read_csv_txt_to_dataframe(spark, filepath, separator)

        return df


class TxtHandler(FileHandlerStrategy):
    def read(self, filepath: str):
        data_out = []
        with open(filepath, "r") as file:
            for line in file:
                data_out.append(line)
        return data_out

    def read_to_dataframe(
        self, spark: SparkSession, filepath: str, separator: str = ";"
    ) -> DataFrame:
        df = read_csv_txt_to_dataframe(spark, filepath, separator)
        return df


class XlsxHandler(FileHandlerStrategy):
    def read(self, filepath):
        dataframe = openpyxl.load_workbook(filepath)
        # Define variable to read sheet
        dataframe1 = dataframe.active

        # Iterate the loop to read the cell values
        sheet_values = []
        for row in range(0, dataframe1.max_row):
            row_values = []
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                row_values.append(col[row].value)
            sheet_values.append(row_values)
        return sheet_values

    def read_to_dataframe(self, spark: SparkSession, filepath: str) -> DataFrame:
        # Solution taken from:
        # https://community.databricks.com/t5/machine-learning/error-when-reading-excel-file-quot-java-lang/td-p/21636
        binary_file = (
            spark.read.format("binaryFile")
            .option("pathGlobFilter", "*.xlsx")
            .load(filepath)
        )

        excel_content = binary_file.head(1)[0].content
        file_like_obj = io.BytesIO(excel_content)

        excel_file = pd.ExcelFile(file_like_obj, engine="openpyxl")
        df_pandas = pd.read_excel(excel_file, dtype=str)

        # Define a schema with all columns as strings
        schema = StructType(
            [StructField(column, StringType(), True) for column in df_pandas.columns]
        )

        # Create the Spark DataFrame using the schema
        df_final = spark.createDataFrame(df_pandas, schema=schema)
        return df_final


class ParquetHandler(FileHandlerStrategy):
    def read(self, filepath: str):
        raise NotImplementedError("Parquet file can't be read to python native object!")

    def read_to_dataframe(self, spark: SparkSession, filepath: str):
        df = spark.read.option("inferSchema", "false").parquet(filepath)

        return df
